import os
import openpyxl
from django.db import models
from django.contrib.staticfiles import finders
from django.utils.crypto import get_random_string
from api.models import CustomUser as User
import secrets
from django.conf import settings


# ------------------------------- Classifications ------------------------------- #


def get_french_cities():
    try:
        french_cities_path = finders.find("data/french_cities.xlsx")
    except Exception as e:
        print("Error getting french cities: ", e)
        french_cities_path = False
    if french_cities_path:
        # File exists, import the cities
        workbook = openpyxl.load_workbook(french_cities_path)
        worksheet = workbook.active
        french_cities = list(set([cell.value for cell in worksheet["A"]]))
    else:
        french_cities = []

    return french_cities


french_cities = get_french_cities()
FRENCH_CITIES_CHOICES = [(city, city) for city in french_cities]


INDUSTRY_CHOICES = [
    ["Administration / Collectivité locale", "Administration / Collectivité locale"],
    [
        "Aéronautique / Marine / Espace / Armement",
        "Aéronautique / Marine / Espace / Armement",
    ],
    ["Agroalimentaire / Agriculture", "Agroalimentaire / Agriculture"],
    ["Alimentation", "Alimentation"],
    ["Associations - ONG", "Associations - ONG"],
    ["Bâtiment", "Bâtiment"],
    ["Biens de consommations", "Biens de consommations"],
    ["Boutique / Magasin", "Boutique / Magasin"],
    ["Communication et médias", "Communication et médias"],
    ["High Tech", "High Tech"],
    ["Conseil / Coaching", "Conseil / Coaching"],
    ["Distribution", "Distribution"],
    ["Enseignement", "Enseignement"],
    ["Finance", "Finance"],
    ["Industrie", "Industrie"],
    ["Marketing de Réseau", "Marketing de Réseau"],
    ["Santé", "Santé"],
    ["Services à domiciles", "Services à domiciles"],
    ["Services aux entreprises", "Services aux entreprises"],
    ["Services de proximité", "Services de proximité"],
    ["Tourisme / Loisirs", "Tourisme / Loisirs"],
    ["Transports", "Transports"],
    ["Gestion des espaces verts", "Gestion des espaces verts"],
    ["Fonctions Support", "Fonctions Support"],
    ["Ressources humaines", "Ressources humaines"],
    ["Gestion de projet", "Gestion de projet"],
    ["Marketing", "Marketing"],
    ["Communication", "Communication"],
    ["Comptabilité et Finance", "Comptabilité et Finance"],
    ["Logistique", "Logistique"],
    ["Administration", "Administration"],
    ["Qualité", "Qualité"],
    ["Autre", "Autre"],
]

CONTRACT_TYPE_CHOICES = [
    ("Contrat d'apprentissage", "Contrat d'apprentissage"),
    ("Contrat de professionnalisation", "Contrat de professionnalisation"),
]

TARGET_EDUCATIONAL_LEVEL_CHOICES = [
    ("Equivalent CAP/BEP", "Equivalent CAP/BEP"),
    ("Bac", "Bac"),
    ("Bac +2", "Bac +2"),
    ("Bac +3", "Bac +3"),
    ("Bac +4", "Bac +4"),
    ("Bac +5", "Bac +5"),
    ("Bac +6 et plus", "Bac +6 et plus"),
]


class Sector(models.Model):
    name = models.CharField(max_length=255, verbose_name="Nom du secteur")

    @classmethod
    def create_from_company_choices(cls):
        print("Creating sectors from company choices")
        for choice in INDUSTRY_CHOICES:
            sector, created = cls.objects.get_or_create(name=choice[0])

    def __str__(self):
        return self.name


class CodeAPE(models.Model):
    code = models.CharField(max_length=5)
    description = models.CharField(max_length=255, editable=False)

    def __str__(self):
        return self.code + " - " + self.description

    @classmethod
    def create_from_excel(cls):
        try:
            # Path to the Excel file in static
            excel_file_path = finders.find("data/codeAPE.xlsx")
            if not excel_file_path:
                raise FileNotFoundError("Excel file not found")
            print("Excel file path: ", excel_file_path)
            # Open the Excel workbook
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active

            # Create objects from the Excel data
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                code, description = row[0], row[1]
                obj, created = cls.objects.get_or_create(
                    code=code, description=description
                )
        except Exception as e:
            print("Error creating CodeAPE objects: ", e)


class Formation(models.Model):
    name = models.CharField(max_length=255, verbose_name="Nom de la formation")
    level = models.CharField(
        choices=TARGET_EDUCATIONAL_LEVEL_CHOICES, max_length=255, default=""
    )

    def __str__(self):
        return self.name + " - " + self.level


# ------------------------------- Companies ------------------------------- #


class TempCompany(models.Model):
    name = models.CharField(max_length=255, default="", blank=True)
    email = models.EmailField(default="", unique=True)
    token = models.CharField(max_length=255, default="", blank=True)
    company = models.OneToOneField(
        "Company", on_delete=models.CASCADE, null=True, blank=True
    )
    date_created = models.DateTimeField(auto_now_add=True, blank=True, null=True)

    def save(self, *args, **kwargs):
        if not self.token:
            self.token = generate_token()
        super().save(*args, **kwargs)

    def __str__(self):
        return self.name

    def link_inscription(self):
        if self.company:
            return None
        else:
            return settings.DOMAIN + f"/register-company/{self.token}"


class Company(models.Model):
    name = models.CharField(
        max_length=255, default="", verbose_name="Nom de l'entreprise"
    )
    sector = models.ManyToManyField(Sector, verbose_name="Secteur d'activité")
    codeAPE = models.ForeignKey(
        CodeAPE, on_delete=models.CASCADE, default="", verbose_name="Code APE"
    )
    adress = models.CharField(max_length=255, default="", verbose_name="Adresse")
    city = models.CharField(max_length=255, default="", verbose_name="Ville")
    zip_code = models.CharField(max_length=20, default="", verbose_name="Code postal")
    country = models.CharField(max_length=255, default="France", verbose_name="Pays")
    description = models.TextField(
        default="", verbose_name="Description de l'activité", blank=True
    )
    created_at = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        return self.name

    def get_contacts(self):
        return [link.user for link in self.company_user_link_set.select_related("user")]


class Skill(models.Model):
    name = models.CharField(max_length=100, default="")

    def __str__(self):
        return self.name


class JobOffer(models.Model):
    company = models.ForeignKey(Company, on_delete=models.CASCADE)
    title = models.CharField(max_length=255, default="")
    same_entity_location = models.BooleanField(default=False)
    location = models.CharField(max_length=255, default="", blank=False)
    contract_type = models.CharField(
        max_length=255, choices=CONTRACT_TYPE_CHOICES, default="Autre"
    )
    target_educational_level = models.CharField(
        choices=TARGET_EDUCATIONAL_LEVEL_CHOICES, max_length=255, default=""
    )
    description = models.TextField(default="")
    skills = models.ManyToManyField(Skill)
    is_active = models.BooleanField(default=False)
    created_at = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        return self.company.name + " - " + self.title


class OfferToken(models.Model):
    offer = models.OneToOneField(JobOffer, on_delete=models.CASCADE)
    token = models.CharField(max_length=20, unique=True)
    created_at = models.DateTimeField(auto_now_add=True)

    def save(self, *args, **kwargs):
        if not self.token:
            self.token = self.generate_token()
        return super().save(*args, **kwargs)

    def generate_token(self):
        return get_random_string(length=20)


# ------------------------------- Applications ------------------------------- #


def get_applicant_resume_path(instance, filename):
    """
    Function to determine the upload path for applicant resumes.
    """
    # Get the file extension
    ext = filename.split(".")[-1]
    # Rename the file to the format: resume_<applicant_id>.<extension>
    new_filename = f"resume_{instance.user.id}_{instance.user.email}.{ext}"
    # Return the full upload path
    return os.path.join("resumes", new_filename)


class Applicant(models.Model):
    user = models.OneToOneField(User, on_delete=models.CASCADE)
    formation = models.ForeignKey(Formation, on_delete=models.CASCADE, null=True)
    resume = models.FileField(
        upload_to=get_applicant_resume_path, verbose_name="CV", blank=True, null=True
    )

    # Informations personnelles

    first_name = models.CharField(max_length=255, default="", verbose_name="Prénom")
    last_name = models.CharField(
        max_length=255, default="", verbose_name="Nom de famille"
    )
    email = models.EmailField(default="", verbose_name="Email")
    city = models.CharField(max_length=255, default="", verbose_name="Ville")
    country = models.CharField(max_length=255, default="", verbose_name="Pays")
    phone = models.CharField(max_length=20, default="", verbose_name="Téléphone")

    # Situation professionnelle actuelle

    diploma = models.CharField(
        max_length=255, default="", verbose_name="Dernier diplôme obtenu"
    )
    # current_company = models.CharField(max_length=255, default="", verbose_name="Entreprise actuelle")

    target_educational_level = models.CharField(
        choices=TARGET_EDUCATIONAL_LEVEL_CHOICES,
        max_length=255,
        default="",
        verbose_name="Niveau de diplôme visé",
    )

    DURATION_CHOICES = [
        ("-1", "Pas de préférence"),
        ("0", "Moins de 6 mois"),
        ("1", "6 mois à 1 an"),
        ("2", "2 ans"),
        ("3", "3 ans"),
    ]
    duration = models.CharField(
        choices=DURATION_CHOICES,
        default="-1",
        verbose_name="Durée du contrat souhaitée",
        max_length=255,
    )
    contract_type = models.CharField(
        choices=CONTRACT_TYPE_CHOICES,
        max_length=255,
        default="",
        verbose_name="Type de contrat souhaité",
    )
    sector = models.ManyToManyField(Sector, verbose_name="Secteur d'activité")

    location = models.CharField(
        max_length=255, default="", verbose_name="Lieu désiré pour le travail"
    )
    kilometers_away = models.IntegerField(
        default=-1,
        verbose_name="Distance maximale (km)",
        choices=[(5, 5), (10, 10), (50, 50), (100, 100), (-1, "Pas de préférence")],
    )

    # Schools
    is_not_signed_in_school = models.BooleanField(
        default=False, verbose_name="Je n'ai pas encore signé avec une école"
    )
    do_accept_school = models.BooleanField(
        default=False,
        verbose_name="Je ne souhaite pas être contacté par les écoles partenaires",
    )

    skills = models.ManyToManyField(Skill, verbose_name="Compétences", blank=True)
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="Créé le")
    updated_at = models.DateTimeField(auto_now=True, verbose_name="Mis à jour le")

    def __str__(self):
        return self.first_name + " " + self.last_name


class TempApplicant(models.Model):
    first_name = models.CharField(max_length=255, default="", blank=True)
    last_name = models.CharField(max_length=255, default="", blank=True)
    email = models.EmailField(default="", unique=True)
    applicant = models.OneToOneField(
        Applicant, on_delete=models.CASCADE, null=True, blank=True
    )
    date_created = models.DateTimeField(auto_now_add=True, blank=True, null=True)
    token = models.CharField(max_length=255, default="", blank=True)

    def save(self, *args, **kwargs):
        if not self.token:  # Generate a token only if it doesn't already exist
            self.token = generate_token()
        super().save(*args, **kwargs)

    def link_inscription(self):
        if self.applicant:
            return None
        else:
            return settings.DOMAIN + f"/register-applicant/{self.token}"


def generate_token():
    return secrets.token_hex(16)


class Application(models.Model):
    applicant = models.OneToOneField(Applicant, on_delete=models.CASCADE)
    created_at = models.DateTimeField(auto_now_add=True)
    is_active = models.BooleanField(default=False)

    def __str__(self):
        return self.applicant.first_name + " " + self.applicant.last_name


# ------------------------------- Matching ------------------------------- #


class match_applicant(models.Model):
    STATUS_CHOICES = [
        ("Pending", "Pending"),
        ("Accepted_company", "Accepted_company"),
        ("Queue_company", "Queue_company"),
        ("Canceled_company", "Canceled_conmapy"),
        ("Canceled_applicant", "Canceled_applicant"),
        ("Fully_accepted", "Fully_accepted"),
        ("Finalized_enrollment", "Finalized_enrollment"),
    ]
    offer = models.ForeignKey(JobOffer, on_delete=models.CASCADE)
    application = models.ForeignKey(Application, on_delete=models.CASCADE)
    score = models.IntegerField(default=0)
    industry_score = models.IntegerField(default=0)
    test_score = models.IntegerField(default=0)
    geographic_score = models.IntegerField(default=0)
    resume_score = models.IntegerField(default=0)
    status = models.CharField(max_length=255, choices=STATUS_CHOICES, default="Pending")
    created_at = models.DateTimeField(auto_now_add=True)


# ------------------------------- Links ------------------------------- #


class company_user_link(models.Model):
    company = models.ForeignKey(Company, on_delete=models.CASCADE)
    user = models.OneToOneField(User, on_delete=models.CASCADE)
    created_at = models.DateTimeField(auto_now_add=True)
    privilege = models.CharField(max_length=255, default="reader")

    def __str__(self):
        return self.user.email + " - " + self.company.name


# ------------------------------- Tests ------------------------------- #
class Application_test(models.Model):
    application = models.OneToOneField(Application, on_delete=models.CASCADE)

    # Section 1: Personnalité
    EXTRAVERSION_CHOICES = [
        ("1", "Pas du tout extraverti"),
        ("2", "Plutôt introverti"),
        ("3", "Modérément extraverti"),
        ("4", "Plutôt extraverti"),
        ("5", "Extrêmement extraverti"),
    ]
    extraversion_verbose = "Je suis"
    extraversion = models.CharField(
        max_length=1, choices=EXTRAVERSION_CHOICES, verbose_name=extraversion_verbose
    )

    TEAMWORK_CHOICES = [
        ("Indépendamment", "Travailler seul"),
        ("En équipe", "Travailler en équipe"),
    ]
    teamwork_preference_verbose = "Ma préférence de travail est"
    teamwork_preference = models.CharField(
        max_length=100,
        choices=TEAMWORK_CHOICES,
        verbose_name=teamwork_preference_verbose,
    )

    REACTION_TO_CHANGE_CHOICES = [
        ("Accueille avec enthousiasme", "Je l'accueille avec enthousiasme"),
        ("Trouve stressantes", "Je le trouve stressant"),
    ]
    reaction_to_change_verbose = "Face au changement"
    reaction_to_change = models.CharField(
        max_length=100,
        choices=REACTION_TO_CHANGE_CHOICES,
        verbose_name=reaction_to_change_verbose,
    )

    DETAIL_ORIENTATION_CHOICES = [
        ("Attentif aux détails", "Orienté vers les détails"),
        ("Vision d'ensemble", "Orienté vers la vision d'ensemble"),
    ]
    detail_orientation_verbose = "Je suis plus orienté vers"
    detail_orientation = models.CharField(
        max_length=100,
        choices=DETAIL_ORIENTATION_CHOICES,
        verbose_name=detail_orientation_verbose,
    )

    PLANNING_PREFERENCE_CHOICES = [
        ("Planifier à l'avance", "Planifier à l'avance"),
        ("Agir de manière spontanée", "Agir de manière spontanée"),
    ]
    planning_preference_verbose = "Je préfère"
    planning_preference = models.CharField(
        max_length=100,
        choices=PLANNING_PREFERENCE_CHOICES,
        verbose_name=planning_preference_verbose,
    )

    # Section 2: Préférences professionnelles
    WORK_ENVIRONMENT_CHOICES = [
        ("Calme et structuré", "Un environnement calme et structuré"),
        ("Dynamique et stimulant", "Un environnement dynamique et stimulant"),
    ]
    work_environment_preference_verbose = "Je préfère travailler dans"
    work_environment_preference = models.CharField(
        max_length=100,
        choices=WORK_ENVIRONMENT_CHOICES,
        verbose_name=work_environment_preference_verbose,
    )

    TASK_PREFERENCE_CHOICES = [
        ("Tâches répétitives", "Des tâches répétitives et bien définies"),
        ("Tâches variées et créatives", "Des tâches variées et créatives"),
    ]
    task_preference_verbose = "Je préfère des tâches"
    task_preference = models.CharField(
        max_length=100,
        choices=TASK_PREFERENCE_CHOICES,
        verbose_name=task_preference_verbose,
    )

    MANAGEMENT_PREFERENCE_CHOICES = [
        ("Supervision étroite", "Une supervision étroite et un suivi régulier"),
        ("Autonomie", "Une autonomie dans la gestion de mes tâches"),
    ]
    management_preference_verbose = "En terme d'organisation, je préfère"
    management_preference = models.CharField(
        max_length=100,
        choices=MANAGEMENT_PREFERENCE_CHOICES,
        verbose_name=management_preference_verbose,
    )

    PRESSURE_REACTION_CHOICES = [
        ("Mieux sous pression", "Je travaille mieux sous pression"),
        (
            "Besoin d'un environnement moins stressant",
            "J'ai besoin d'un environnement moins stressant pour bien performer",
        ),
    ]
    pressure_reaction_verbose = "En termes de pression :"
    pressure_reaction = models.CharField(
        max_length=100,
        choices=PRESSURE_REACTION_CHOICES,
        verbose_name=pressure_reaction_verbose,
    )

    WORK_MOTIVATION_CHOICES = [
        ("Défis et apprentissage", "Les défis et les possibilités d'apprentissage"),
        ("Stabilité et sécurité", "La stabilité et la sécurité d'emploi"),
    ]
    work_motivation_verbose = "Ma motivation au travail est"
    work_motivation = models.CharField(
        max_length=100,
        choices=WORK_MOTIVATION_CHOICES,
        verbose_name=work_motivation_verbose,
    )

    # Section 3: Intérêts et passions
    INTERACTION_PREFERENCE_CHOICES = [
        ("Interactions humaines", "Les interactions avec les autres"),
        ("Problèmes techniques", "La résolution de problèmes techniques"),
    ]
    interaction_preference_verbose = "Au sein de l'entreprise, je préfère :"
    interaction_preference = models.CharField(
        max_length=100,
        choices=INTERACTION_PREFERENCE_CHOICES,
        verbose_name=interaction_preference_verbose,
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def __str__(self):
        return f"Application_test - id: {self.id}"


# ---------------------- Offer_test ---------------------- #


class Offer_test(models.Model):
    offer = models.OneToOneField(JobOffer, on_delete=models.CASCADE)

    # Section 1: Traits de personnalité
    EXTRAVERSION_CHOICES = [
        ("1", "Pas du tout nécessaire"),
        ("2", "Plutôt peu nécessaire"),
        ("3", "Modérément nécessaire"),
        ("4", "Plutôt nécessaire"),
        ("5", "Extrêmement nécessaire"),
    ]
    extraversion = models.CharField(
        max_length=1,
        choices=EXTRAVERSION_CHOICES,
        verbose_name="Niveau d'extraversion requis",
    )

    TEAMWORK_CHOICES = [
        ("Indépendamment", "Faible importance"),
        ("En équipe", "Grande importance"),
    ]
    teamwork_preference = models.CharField(
        max_length=100,
        choices=TEAMWORK_CHOICES,
        verbose_name="Importance du travail en équipe",
    )

    REACTION_TO_CHANGE_CHOICES = [
        ("Accueille avec enthousiasme", "Accueille les changements avec enthousiasme"),
        ("Trouve stressantes", "Trouve les changements stressants"),
    ]
    reaction_to_change = models.CharField(
        max_length=100,
        choices=REACTION_TO_CHANGE_CHOICES,
        verbose_name="Réaction face au changement",
    )

    DETAIL_ORIENTATION_CHOICES = [
        ("Attentif aux détails", "Attentif aux détails"),
        ("Vision d'ensemble", "Vision d'ensemble"),
    ]
    detail_orientation = models.CharField(
        max_length=100,
        choices=DETAIL_ORIENTATION_CHOICES,
        verbose_name="Sens du détail",
    )

    PLANNING_PREFERENCE_CHOICES = [
        ("Planifier à l'avance", "Planification minutieuse à l'avance"),
        ("Agir de manière spontanée", "Flexibilité et adaptabilité"),
    ]
    planning_preference = models.CharField(
        max_length=100,
        choices=PLANNING_PREFERENCE_CHOICES,
        verbose_name="Gestion du temps",
    )

    # Section 2: Préférences professionnelles
    WORK_ENVIRONMENT_CHOICES = [
        ("Calme et structuré", "Un environnement calme et structuré"),
        ("Dynamique et stimulant", "Un environnement dynamique et stimulant"),
    ]
    work_environment_preference = models.CharField(
        max_length=100,
        choices=WORK_ENVIRONMENT_CHOICES,
        verbose_name="Environnement de travail préféré",
    )

    TASK_PREFERENCE_CHOICES = [
        ("Tâches répétitives", "Tâches répétitives et bien définies"),
        ("Tâches variées et créatives", "Tâches variées et créatives"),
    ]
    task_preference = models.CharField(
        max_length=100,
        choices=TASK_PREFERENCE_CHOICES,
        verbose_name="Type de tâches préféré",
    )

    MANAGEMENT_PREFERENCE_CHOICES = [
        ("Supervision étroite", "Supervision étroite avec suivi régulier"),
        ("Autonomie", "Autonomie dans la gestion des tâches"),
    ]
    management_preference = models.CharField(
        max_length=100,
        choices=MANAGEMENT_PREFERENCE_CHOICES,
        verbose_name="Style de supervision préféré",
    )

    PRESSURE_REACTION_CHOICES = [
        ("Mieux sous pression", "Travaille mieux sous pression"),
        (
            "Besoin d'un environnement moins stressant",
            "Besoin d'un environnement moins stressant pour bien performer",
        ),
    ]
    pressure_reaction = models.CharField(
        max_length=100,
        choices=PRESSURE_REACTION_CHOICES,
        verbose_name="Réaction sous pression",
    )

    WORK_MOTIVATION_CHOICES = [
        ("Défis et apprentissage", "Les défis et les possibilités d'apprentissage"),
        ("Stabilité et sécurité", "La stabilité et la sécurité d'emploi"),
    ]
    work_motivation = models.CharField(
        max_length=100,
        choices=WORK_MOTIVATION_CHOICES,
        verbose_name="Facteurs de motivation",
    )

    # Section 3: Attentes spécifiques de l'entreprise
    INTERACTION_PREFERENCE_CHOICES = [
        ("Interactions humaines", "Interactions humaines"),
        ("Problèmes techniques", "Résolution de problèmes techniques"),
    ]
    interaction_preference = models.CharField(
        max_length=100,
        choices=INTERACTION_PREFERENCE_CHOICES,
        verbose_name="Préférence d'interaction",
    )

    def __str__(self):
        return f"Offer_test - id: {self.id}"
